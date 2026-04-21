"""
extract/incasol.py
Extracción de los 4 Excel trimestrales de Incasol para barrios y distritos.
Fuente: Generalitat de Catalunya · 2000-2025
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.config import PipelineConfig


class IncasolExtractor:
    """
    Parsea los 4 Excel Incasol (contratos, alquiler, alquiler/m², superficie)
    y construye los paneles brutos de barrios y distritos.

    Uso:
        extractor = IncasolExtractor(config)
        barris_raw   = extractor.extract_barris()    # 3504 × 8
        district_raw = extractor.extract_districts() # 1040 × 8
    """

    _NULL_STRINGS = {"n.d.", "nd", "n.d", "na", "n/a", "-", ""}

    _EXCEL_FILES = {
        "num_contracts": ("trimestral_bcn_contractes.xlsx", True),   # (filename, as_int)
        "avg_rent":      ("trimestral_bcn_lloguer.xlsx",    False),
        "avg_rent_m2":   ("trimestral_bcn_lloguer_m2.xlsx", False),
        "avg_surface":   ("trimestral_bcn_sup.xlsx",        False),
    }

    def __init__(self, config: PipelineConfig):
        self.raw_path = config.raw_path
        self._barris_start  = config.barris_start_year
        self._barris_end    = config.barris_end_year
        self._district_start = config.district_start_year
        self._district_end   = config.district_end_year

    # ── Helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _clean_text(value) -> str:
        if value is None:
            return ""
        return " ".join(str(value).strip().split())

    def _parse_number(self, value, as_int: bool = False):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return int(value) if as_int else float(value)
        text = str(value).strip().replace(",", ".")
        if text.lower() in self._NULL_STRINGS:
            return None
        try:
            return int(float(text)) if as_int else float(text)
        except ValueError:
            return None

    def _year_sheets(self, filepath: Path, start: int, end: int) -> list:
        wb = load_workbook(filepath, data_only=True)
        return sorted(
            [s for s in wb.sheetnames if s.isdigit() and start <= int(s) <= end],
            key=int,
        )

    # ── Parseo de barrios ────────────────────────────────────────────────────

    def _parse_sheet_barris(self, filepath: Path, sheet: str, col: str, as_int: bool) -> list:
        wb   = load_workbook(filepath, data_only=True)
        rows = list(wb[sheet].iter_rows(values_only=True))
        start = next(
            (i for i, r in enumerate(rows)
             if self._clean_text(r[1] if len(r) > 1 else "").startswith("Barris")),
            None,
        )
        if start is None:
            raise ValueError(f"Sección 'Barris' no encontrada en hoja {sheet}")
        year, records = int(sheet), []
        for row in rows[start + 1:]:
            code = row[0] if row else None
            name = self._clean_text(row[1]) if len(row) > 1 else ""
            if not isinstance(code, int) or not name:
                break
            for col_idx, quarter in zip([2, 3, 4, 5], [1, 2, 3, 4]):
                records.append({
                    "neighborhood_code": code,
                    "neighborhood": name,
                    "year": year,
                    "quarter": quarter,
                    col: self._parse_number(row[col_idx] if col_idx < len(row) else None, as_int),
                })
        return records

    def _extract_one_barris(self, col: str, filename: str, as_int: bool) -> pd.DataFrame:
        path   = self.raw_path / filename
        sheets = self._year_sheets(path, self._barris_start, self._barris_end)
        rows   = []
        for s in sheets:
            rows.extend(self._parse_sheet_barris(path, s, col, as_int))
        return (
            pd.DataFrame(rows)
            .sort_values(["year", "neighborhood_code", "quarter"])
            .reset_index(drop=True)
        )

    def extract_barris(self) -> pd.DataFrame:
        """Extrae y mergea los 4 Excel Incasol de barrios. Retorna 3504 × 8."""
        frames = {
            col: self._extract_one_barris(col, fname, as_int)
            for col, (fname, as_int) in self._EXCEL_FILES.items()
        }
        key = ["neighborhood_code", "year", "quarter"]
        df  = frames["num_contracts"]
        for col in ["avg_rent", "avg_rent_m2", "avg_surface"]:
            df = df.merge(frames[col][key + [col]], on=key, how="left", validate="1:1")
        df["num_contracts"] = df["num_contracts"].astype("Int64")
        return df[["neighborhood_code", "neighborhood", "year", "quarter",
                   "num_contracts", "avg_rent", "avg_rent_m2", "avg_surface"]]

    # ── Parseo de distritos ──────────────────────────────────────────────────

    def _parse_sheet_districts_new(self, rows: list, year: int, col: str, as_int: bool) -> list:
        start = next(
            (i for i, r in enumerate(rows)
             if self._clean_text(r[1] if len(r) > 1 else "").startswith("Districtes municipals")),
            None,
        )
        if start is None:
            raise ValueError(f"Sección 'Districtes municipals' no encontrada para {year}")
        records = []
        for row in rows[start + 1:]:
            code = row[0] if row else None
            name = self._clean_text(row[1]) if len(row) > 1 else ""
            if not isinstance(code, int) or not name:
                break
            for col_idx, quarter in zip([2, 3, 4, 5], [1, 2, 3, 4]):
                records.append({
                    "district_code": code, "district": name,
                    "year": year, "quarter": quarter,
                    col: self._parse_number(row[col_idx] if col_idx < len(row) else None, as_int),
                })
        return records

    def _parse_sheet_districts_old(self, rows: list, year: int, col: str, as_int: bool) -> list:
        records = []
        for row in rows:
            first = self._clean_text(row[0]) if row else ""
            if not first or first == "Barcelona":
                continue
            m = re.match(r"^(\d+)\.\s*(.+)$", first)
            if not m:
                continue
            code, name = int(m.group(1)), m.group(2).strip()
            for col_idx, quarter in zip([1, 2, 3, 4], [1, 2, 3, 4]):
                records.append({
                    "district_code": code, "district": name,
                    "year": year, "quarter": quarter,
                    col: self._parse_number(row[col_idx] if col_idx < len(row) else None, as_int),
                })
        return records

    def _extract_one_districts(self, col: str, filename: str, as_int: bool) -> pd.DataFrame:
        path   = self.raw_path / filename
        sheets = self._year_sheets(path, self._district_start, self._district_end)
        rows_all = []
        for s in sheets:
            wb   = load_workbook(path, data_only=True)
            rows = list(wb[s].iter_rows(values_only=True))
            year = int(s)
            fn   = self._parse_sheet_districts_new if year >= 2014 else self._parse_sheet_districts_old
            rows_all.extend(fn(rows, year, col, as_int))
        return (
            pd.DataFrame(rows_all)
            .sort_values(["year", "district_code", "quarter"])
            .reset_index(drop=True)
        )

    def extract_districts(self) -> pd.DataFrame:
        """Extrae y mergea los 4 Excel Incasol de distritos. Retorna 1040 × 8."""
        frames = {
            col: self._extract_one_districts(col, fname, as_int)
            for col, (fname, as_int) in self._EXCEL_FILES.items()
        }
        key = ["district_code", "year", "quarter"]
        df  = frames["num_contracts"]
        for col in ["avg_rent", "avg_rent_m2", "avg_surface"]:
            df = df.merge(frames[col][key + [col]], on=key, how="left", validate="1:1")
        df["num_contracts"] = df["num_contracts"].astype("Int64")
        return df[["district_code", "district", "year", "quarter",
                   "num_contracts", "avg_rent", "avg_rent_m2", "avg_surface"]]
